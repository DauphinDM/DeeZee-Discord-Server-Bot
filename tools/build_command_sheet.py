"""
Generate the merged Deezee Server Bot command sheet as an .xlsx workbook.

The sheet merges the public command surfaces of Dyno, Carl-bot, Lawliet,
Sapphire and Double Counter into one deduplicated command set for a single
replacement bot using the "?" prefix.

Run:  python tools/build_command_sheet.py
Out:  docs/Deezee_Command_Sheet.xlsx
"""

from __future__ import annotations

import os
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADERS = [
    "Command",
    "Category",
    "Aliases",
    "Slash / Prefix / Both",
    "Description",
    "Arguments",
    "Required permissions",
    "Source bot(s)",
    "Uses buttons?",
    "Uses modal?",
    "Notes",
]

# Row = (command, category, aliases, surface, description, arguments,
#        permissions, sources, buttons, modal, notes)
Row = Tuple[str, str, str, str, str, str, str, str, str, str, str]

MODERATION: List[Row] = [
    ("?ban", "Moderation", "-", "Both",
     "Ban a member or a raw user ID, optionally for a fixed duration, with optional message deletion.",
     "<user|user_id> [duration] [reason]", "Ban Members",
     "Dyno, Carl-bot, Lawliet, Sapphire", "Yes", "No",
     "Confirm/cancel buttons locked to the invoker. Duration accepts 10m/2h/7d. Timed bans stored in the DB and lifted by a scheduler task."),
    ("?unban", "Moderation", "pardon", "Both",
     "Lift an active ban and close the matching mod-log case.",
     "<user_id|user_tag> [reason]", "Ban Members",
     "Dyno, Carl-bot, Lawliet", "No", "No",
     "Autocomplete on the slash version lists the guild's current bans."),
    ("?softban", "Moderation", "-", "Both",
     "Ban then immediately unban a member to bulk-delete their recent messages.",
     "<user> [delete_days=1] [reason]", "Ban Members",
     "Dyno, Carl-bot", "Yes", "No",
     "Confirmation buttons. Useful for cleaning up spam without a permanent ban."),
    ("?massban", "Moderation", "banlist", "Both",
     "Ban a pasted list of user IDs in one action, with a progress report.",
     "<reason>", "Ban Members + Manage Guild",
     "Dyno, Carl-bot", "Yes", "Yes",
     "Modal takes the newline-separated ID list so the command line stays short. Rate-limit aware: batched with backoff."),
    ("?kick", "Moderation", "-", "Both",
     "Remove a member from the server; they can rejoin with a new invite.",
     "<user> [reason]", "Kick Members",
     "Dyno, Carl-bot, Lawliet, Sapphire", "Yes", "No",
     "Confirmation buttons. DM notice sent before the kick, failures reported but non-fatal."),
    ("?mute", "Moderation", "timeout, tempmute", "Both",
     "Time out a member so they cannot send messages, react or speak.",
     "<user> [duration=1h] [reason]", "Moderate Members",
     "Dyno, Carl-bot, Lawliet, Sapphire", "Yes", "No",
     "Uses Discord's native timeout up to 28 days; falls back to a configured mute role for longer durations."),
    ("?unmute", "Moderation", "untimeout", "Both",
     "Clear an active timeout or remove the mute role.",
     "<user> [reason]", "Moderate Members",
     "Dyno, Carl-bot, Lawliet, Sapphire", "No", "No", "-"),
    ("?warn", "Moderation", "-", "Both",
     "Issue a formal warning, log a case and DM the member.",
     "<user> <reason>", "Moderate Members",
     "Dyno, Carl-bot, Lawliet", "No", "No",
     "Feeds the automod escalation ladder: the configured warn threshold can auto-mute/kick/ban."),
    ("?warnings", "Moderation", "warns, infractions", "Both",
     "List a member's warnings with case IDs, issuing moderator and timestamps.",
     "[user] [page]", "Moderate Members",
     "Dyno, Carl-bot, Lawliet", "Yes", "No",
     "Paginated embed with previous/next/jump buttons."),
    ("?delwarn", "Moderation", "removewarn", "Both",
     "Delete a single warning by case ID.",
     "<case_id> [reason]", "Moderate Members", "Dyno, Carl-bot", "No", "No", "-"),
    ("?clearwarns", "Moderation", "clearwarn", "Both",
     "Delete every warning on a member's record.",
     "<user>", "Manage Guild", "Dyno, Carl-bot", "Yes", "No",
     "Destructive: confirm/cancel buttons required."),
    ("?case", "Moderation", "-", "Both",
     "Show a single mod-log case in full.",
     "<case_id>", "Moderate Members", "Dyno, Carl-bot", "No", "No", "-"),
    ("?reason", "Moderation", "editcase", "Both",
     "Edit the reason on an existing case and update the posted mod-log embed.",
     "<case_id> <new_reason>", "Moderate Members", "Dyno, Carl-bot", "No", "Yes",
     "Modal used for the slash version so long reasons are easy to type."),
    ("?modlogs", "Moderation", "history", "Both",
     "Show every case tied to a member across all action types.",
     "<user> [page]", "Moderate Members", "Dyno, Carl-bot", "Yes", "No",
     "Paginated. Filter buttons per action type (ban/kick/mute/warn)."),
    ("?moderations", "Moderation", "activemods", "Both",
     "List punishments that are still active and scheduled to expire.",
     "[user] [page]", "Moderate Members", "Dyno", "Yes", "No", "-"),
    ("?duration", "Moderation", "-", "Both",
     "Change the remaining duration of an active mute or timed ban.",
     "<case_id> <duration>", "Moderate Members", "Dyno", "No", "No", "-"),
    ("?modstats", "Moderation", "-", "Both",
     "Show per-moderator action counts over a time window.",
     "[moderator] [days=30]", "Manage Guild", "Dyno", "No", "No", "-"),
    ("?note", "Moderation", "addnote", "Both",
     "Attach a private staff note to a member. Notes never DM the member.",
     "<user> <text>", "Moderate Members", "Dyno", "No", "Yes", "-"),
    ("?notes", "Moderation", "-", "Both",
     "List private staff notes on a member.",
     "<user>", "Moderate Members", "Dyno", "Yes", "No", "Paginated."),
    ("?delnote", "Moderation", "-", "Both",
     "Delete a single staff note by its ID.",
     "<user> <note_id>", "Moderate Members", "Dyno", "No", "No", "-"),
    ("?clearnotes", "Moderation", "-", "Both",
     "Delete every staff note on a member.",
     "<user>", "Manage Guild", "Dyno", "Yes", "No", "Confirmation buttons."),
    ("?purge", "Moderation", "clear, prune", "Both",
     "Bulk-delete recent messages, optionally filtered by author, content or attachment type.",
     "<count 1-1000> [user] [--bots|--humans|--links|--images|--embeds|--contains <text>]",
     "Manage Messages", "Dyno, Carl-bot, Sapphire", "Yes", "No",
     "Confirmation buttons above 100 messages. Discord cannot bulk-delete messages older than 14 days; those are skipped and reported."),
    ("?slowmode", "Moderation", "sm", "Both",
     "Set or clear per-channel slowmode.",
     "[channel] <seconds|off>", "Manage Channels", "Dyno, Carl-bot", "No", "No", "-"),
    ("?lock", "Moderation", "-", "Both",
     "Deny Send Messages to @everyone in a channel, optionally for a fixed time.",
     "[channel] [duration] [reason]", "Manage Channels", "Dyno, Carl-bot", "No", "No",
     "Stores the previous overwrite so ?unlock restores it exactly."),
    ("?unlock", "Moderation", "-", "Both",
     "Restore a locked channel's original permission overwrites.",
     "[channel]", "Manage Channels", "Dyno, Carl-bot", "No", "No", "-"),
    ("?lockdown", "Moderation", "-", "Both",
     "Lock or unlock every channel in a configured lockdown set at once.",
     "<start|end> [reason]", "Manage Guild", "Dyno", "Yes", "No",
     "Confirmation buttons. The channel set is chosen in ?config > Moderation."),
    ("?nuke", "Moderation", "clone", "Both",
     "Clone a channel and delete the original, wiping all history while keeping settings and position.",
     "[channel]", "Manage Channels", "Carl-bot", "Yes", "No",
     "Irreversible: double confirmation, invoker-locked, 30s timeout."),
    ("?nickname", "Moderation", "nick, setnick", "Both",
     "Change or clear a member's nickname.",
     "<user> [nickname]", "Manage Nicknames", "Dyno, Carl-bot", "No", "No", "-"),
    ("?voicekick", "Moderation", "vckick", "Both",
     "Disconnect a member from their voice channel.",
     "<user> [reason]", "Move Members", "Dyno", "No", "No", "-"),
    ("?voicemute", "Moderation", "vcmute", "Both",
     "Server-mute a member in voice channels.",
     "<user> [reason]", "Mute Members", "Dyno", "No", "No", "-"),
    ("?deafen", "Moderation", "undeafen", "Both",
     "Server-deafen or undeafen a member in voice channels.",
     "<user> [reason]", "Deafen Members", "Dyno", "No", "No",
     "One command, toggles based on current state."),
    ("?banmessage", "Moderation", "-", "Both",
     "Set the DM sent to a member when they are banned (supports appeal link and variables).",
     "[message]", "Manage Guild", "Carl-bot", "Yes", "Yes",
     "Modal for the message body; preview button before saving."),
]

AUTOMOD: List[Row] = [
    ("?automod", "Automod", "am", "Both",
     "Open the automod control panel: enable filters, set punishments, manage exemptions.",
     "-", "Manage Guild", "Dyno, Carl-bot, Sapphire", "Yes", "Yes",
     "Root panel. Category buttons open sub-panels; select menus pick roles/channels; modals take numeric thresholds."),
    ("?filter words", "Automod", "censor, bannedwords", "Both",
     "Manage the banned-word list, including wildcard and whole-word matching modes.",
     "<add|remove|list|clear> [words...]", "Manage Guild", "Dyno, Carl-bot", "Yes", "Yes",
     "Modal takes a comma-separated bulk list. Matching normalises leetspeak and zero-width characters."),
    ("?filter invites", "Automod", "-", "Both",
     "Block Discord invite links, with an allowlist of permitted guild IDs.",
     "<on|off|allow|deny> [guild_id]", "Manage Guild", "Dyno, Carl-bot", "Yes", "No", "-"),
    ("?filter links", "Automod", "-", "Both",
     "Block all links, or run an allowlist/blocklist of domains.",
     "<on|off|mode|allow|deny> [domain]", "Manage Guild", "Dyno, Carl-bot", "Yes", "No", "-"),
    ("?filter caps", "Automod", "-", "Both",
     "Delete messages that exceed a percentage of uppercase characters.",
     "<on|off|percent> [1-100] [min_length]", "Manage Guild", "Dyno, Carl-bot", "Yes", "Yes", "-"),
    ("?filter spam", "Automod", "-", "Both",
     "Rate-limit messages per user: N messages in M seconds triggers the configured punishment.",
     "<on|off|rate> [count] [seconds]", "Manage Guild", "Dyno, Carl-bot, Sapphire", "Yes", "Yes", "-"),
    ("?filter mentions", "Automod", "massmention", "Both",
     "Punish messages containing more than N user or role mentions.",
     "<on|off|limit> [count]", "Manage Guild", "Dyno, Carl-bot", "Yes", "Yes", "-"),
    ("?filter emoji", "Automod", "-", "Both",
     "Punish messages containing more than N emoji.",
     "<on|off|limit> [count]", "Manage Guild", "Dyno, Carl-bot", "Yes", "Yes", "-"),
    ("?filter attachments", "Automod", "attachmentspam", "Both",
     "Rate-limit attachments and optionally block specific file extensions.",
     "<on|off|rate|extensions> [args]", "Manage Guild", "Carl-bot", "Yes", "Yes",
     "Extension blocklist defaults to executable and script types."),
    ("?filter zalgo", "Automod", "-", "Both",
     "Delete messages using excessive combining characters (zalgo text).",
     "<on|off>", "Manage Guild", "Dyno, Carl-bot", "Yes", "No", "-"),
    ("?filter duplicates", "Automod", "repeat", "Both",
     "Punish the same message repeated N times, including across channels.",
     "<on|off|limit> [count]", "Manage Guild", "Dyno", "Yes", "Yes", "-"),
    ("?filter newlines", "Automod", "wall", "Both",
     "Punish wall-of-text messages above a newline or character count.",
     "<on|off|limit> [count]", "Manage Guild", "Dyno", "Yes", "Yes", "-"),
    ("?filter stickers", "Automod", "-", "Both",
     "Block stickers, or rate-limit them like attachments.",
     "<on|off|rate> [count] [seconds]", "Manage Guild", "Carl-bot", "Yes", "No", "-"),
    ("?scanlinks", "Automod", "linkscan", "Both",
     "Scan only unknown links. URLs whose domain is on the trusted list are allowed without any lookup; everything else is checked against Google Safe Browsing.",
     "<on|off|action|status|trust|untrust|trustlist> [domain]", "Manage Guild",
     "Dyno (premium), Carl-bot", "Yes", "Yes",
     "EXTERNAL API REQUIRED: Google Safe Browsing v5 urls:search (free, non-commercial). Three-tier check: (1) bundled trusted-domain database shipped as data/trusted_domains.txt; (2) per-guild trusted additions via ?scanlinks trust; (3) only URLs matching neither tier hit the API, batched up to 50 per request. Verdicts cached in SQLite for the cacheDuration the API returns (300s when tested), so a link posted repeatedly inside that window is scanned once. Without a key, tiers 1 and 2 still run and unknown links fall back to a bundled phishing-domain blocklist. Targeting v5, not v4: v4 is deprecated and shuts down 31 March 2027."),
    ("?automod whitelist", "Automod", "amexempt", "Both",
     "Exempt roles, channels or members from all or specific automod filters.",
     "<add|remove|list> <role|channel|member>", "Manage Guild", "Carl-bot", "Yes", "No",
     "Role/channel select menus instead of raw IDs."),
    ("?automod punishment", "Automod", "-", "Both",
     "Set the punishment ladder applied when a filter or the warn threshold trips.",
     "<filter> <delete|warn|mute|kick|ban> [duration]", "Manage Guild",
     "Dyno, Carl-bot, Sapphire", "Yes", "Yes", "-"),
    ("?automod threshold", "Automod", "-", "Both",
     "Set how many warnings within a time window trigger the escalation punishment.",
     "<count> [window]", "Manage Guild", "Carl-bot", "No", "Yes", "-"),
    ("?mediaonly", "Automod", "mediachannel", "Both",
     "Restrict a channel to messages containing an attachment or link.",
     "<add|remove|list> [channel]", "Manage Guild", "Carl-bot", "Yes", "No", "-"),
    ("?automod test", "Automod", "-", "Both",
     "Run a sample message through the active filters and report which would trip, without punishing.",
     "<message>", "Manage Guild", "Deezee original", "No", "Yes",
     "Not in any source bot; added because configuring filters blind is the main support burden."),
]

ANTIRAID: List[Row] = [
    ("?antiraid", "Anti-raid & Verification", "raid", "Both",
     "Open the anti-raid panel: join-rate thresholds, automatic action, alert channel.",
     "-", "Manage Guild", "Dyno, Carl-bot, Double Counter", "Yes", "Yes", "-"),
    ("?raidmode", "Anti-raid & Verification", "-", "Both",
     "Manually force raid mode on or off (gate joins, restrict invites, alert staff).",
     "<on|off> [reason]", "Manage Guild", "Dyno, Double Counter", "Yes", "No",
     "Auto-enables when the join-rate threshold trips; auto-disables after a quiet period."),
    ("?panic", "Anti-raid & Verification", "emergency", "Both",
     "Emergency response: lock every channel, pause invites and quarantine members who joined in the last N minutes.",
     "[minutes=10]", "Administrator", "Double Counter", "Yes", "No",
     "Double confirmation, invoker-locked. Shows a dry-run count before acting."),
    ("?verification", "Anti-raid & Verification", "verifyconfig", "Both",
     "Configure the verification gate: mode, verified role, unverified role, captcha type, timeout.",
     "-", "Manage Guild", "Double Counter, Carl-bot", "Yes", "Yes",
     "Modes: off / on-click (button in a channel) / on-join (DM or gate channel). Persistent view survives restarts."),
    ("?verify", "Anti-raid & Verification", "-", "Both",
     "Member-facing command to start or retry verification.",
     "-", "None", "Double Counter", "Yes", "Yes",
     "Button opens a modal containing a generated image/text captcha. No external service; captcha rendered locally with Pillow."),
    ("?forceverify", "Anti-raid & Verification", "-", "Both",
     "Manually mark a member verified, or revoke their verification.",
     "<user> [revoke]", "Manage Roles", "Double Counter", "No", "No", "-"),
    ("?altcheck", "Anti-raid & Verification", "risk", "Both",
     "Score a member's alt-account risk from account age, avatar, name similarity to existing members, join-burst clustering and invite source.",
     "<user>", "Moderate Members", "Double Counter", "Yes", "No",
     "HEURISTIC ONLY. Double Counter's real detection uses IP/cookie fingerprinting via a hosted verification page; that is out of scope for a self-hosted bot (see Notes sheet)."),
    ("?altconfig", "Anti-raid & Verification", "-", "Both",
     "Set the alt-risk score thresholds and the automatic action taken at each threshold.",
     "-", "Manage Guild", "Double Counter", "Yes", "Yes", "-"),
    ("?minage", "Anti-raid & Verification", "accountage", "Both",
     "Reject or quarantine joins from accounts younger than a set age.",
     "<duration|off> [kick|ban|quarantine]", "Manage Guild", "Dyno, Double Counter", "Yes", "No", "-"),
    ("?joins", "Anti-raid & Verification", "joinwatch, recentjoins", "Both",
     "List recent joins with account age, invite used and computed risk score.",
     "[minutes=60]", "Moderate Members", "Double Counter", "Yes", "No",
     "Paginated. Select members and mass-kick/ban straight from the panel."),
    ("?flagged", "Anti-raid & Verification", "altlist", "Both",
     "List members currently flagged by the alt heuristics, with their score breakdown.",
     "[page]", "Moderate Members", "Double Counter", "Yes", "No", "-"),
    ("?quarantine", "Anti-raid & Verification", "-", "Both",
     "Strip a member's roles and apply the quarantine role pending staff review.",
     "<user> [reason]", "Manage Roles", "Double Counter", "No", "No",
     "Original roles stored so ?unquarantine restores them."),
    ("?unquarantine", "Anti-raid & Verification", "-", "Both",
     "Release a member from quarantine and restore their previous roles.",
     "<user>", "Manage Roles", "Double Counter", "No", "No", "-"),
    ("?invitetrack", "Anti-raid & Verification", "invites", "Both",
     "Show which invite a member joined on, and per-inviter join counts.",
     "[user]", "Manage Guild", "Lawliet, Double Counter", "Yes", "No",
     "Requires the Manage Guild permission on the bot to read invite uses."),
    ("(not built) VPN / proxy / Tor blocking", "Anti-raid & Verification", "-", "n/a",
     "Block joins originating from VPNs, proxies or Tor exit nodes.", "-", "-", "Double Counter", "No", "No",
     "NOT BUILT. Discord never exposes member IP addresses to bots; Double Counter obtains them by redirecting users to its own web page. That requires hosting a public site, which the no-dashboard constraint rules out."),
]

ROLES: List[Row] = [
    ("?role", "Roles & Reaction Roles", "r", "Both",
     "Add or remove a role from a member; toggles if no action is given.",
     "<user> <role> [add|remove]", "Manage Roles",
     "Dyno, Carl-bot, Sapphire", "No", "No",
     "Hierarchy-checked against both the invoker's top role and the bot's."),
    ("?massrole", "Roles & Reaction Roles", "roleall", "Both",
     "Add or remove a role across a whole population: everyone, humans, bots, or holders of another role.",
     "<add|remove> <role> [everyone|humans|bots|in:<role>]", "Manage Guild",
     "Dyno, Carl-bot", "Yes", "No",
     "Confirmation buttons showing the affected count. Runs as a throttled background job with a live progress embed."),
    ("?temprole", "Roles & Reaction Roles", "-", "Both",
     "Give a member a role that is automatically removed after a duration.",
     "<user> <role> <duration> [reason]", "Manage Roles", "Dyno, Carl-bot", "No", "No", "-"),
    ("?createrole", "Roles & Reaction Roles", "addrole", "Both",
     "Create a new guild role with optional colour and hoist.",
     "<name> [color] [hoist]", "Manage Roles", "Dyno", "No", "Yes", "-"),
    ("?deleterole", "Roles & Reaction Roles", "delrole", "Both",
     "Delete a guild role.", "<role>", "Manage Roles", "Dyno", "Yes", "No", "Confirmation buttons."),
    ("?rolecolor", "Roles & Reaction Roles", "-", "Both",
     "Change a role's colour.", "<role> <hex|random>", "Manage Roles", "Dyno", "No", "No", "-"),
    ("?mentionable", "Roles & Reaction Roles", "-", "Both",
     "Toggle whether a role can be mentioned.", "<role> [on|off]", "Manage Roles", "Dyno", "No", "No", "-"),
    ("?roleinfo", "Roles & Reaction Roles", "-", "Both",
     "Show a role's ID, colour, permissions, member count and position.",
     "<role>", "None", "Dyno, Carl-bot", "No", "No", "-"),
    ("?roles", "Roles & Reaction Roles", "rolelist", "Both",
     "List every role with member counts, or list the members holding one role.",
     "[role] [page]", "None", "Dyno", "Yes", "No", "Paginated."),
    ("?reactionrole", "Roles & Reaction Roles", "rr, rolemenu", "Both",
     "Build and manage self-assign role menus using reactions, buttons or a select menu.",
     "<create|add|remove|list|edit|delete> [args]", "Manage Roles",
     "Carl-bot, Sapphire, Lawliet", "Yes", "Yes",
     "Full builder flow: modal for embed text, select menus to pick roles, buttons to choose menu type and limits. Button/select menus registered with bot.add_view() on ready so they keep working after a restart."),
    ("?rr mode", "Roles & Reaction Roles", "-", "Both",
     "Set a menu's behaviour: normal, unique (one role only), verify (add only), drop (remove only), or a max-roles limit.",
     "<menu_id> <normal|unique|verify|drop|limit> [n]", "Manage Roles",
     "Carl-bot", "Yes", "No", "-"),
    ("?autorole", "Roles & Reaction Roles", "-", "Both",
     "Manage roles automatically granted to members (and to bots) on join.",
     "<add|remove|list> [role] [--bots] [--delay <duration>]", "Manage Roles",
     "Carl-bot, Dyno, Sapphire", "Yes", "No",
     "Optional delay defers the grant, which blunts join-and-spam raids."),
    ("?stickyroles", "Roles & Reaction Roles", "-", "Both",
     "Restore a member's previous roles when they rejoin the server.",
     "<on|off|blacklist> [role]", "Manage Roles", "Carl-bot", "Yes", "No", "-"),
    ("?selfrole", "Roles & Reaction Roles", "rank, iam", "Both",
     "Join or leave a role that has been marked self-assignable.",
     "[role]", "None", "Dyno", "Yes", "No",
     "With no argument, shows a select menu of every available self-role."),
    ("?selfroles", "Roles & Reaction Roles", "ranks", "Both",
     "Manage which roles members may self-assign.",
     "<add|remove|list> [role]", "Manage Roles", "Dyno", "Yes", "No", "-"),
]

LOGGING: List[Row] = [
    ("?logging", "Logging", "logs, log", "Both",
     "Open the logging panel: pick a destination channel per event group and toggle individual events.",
     "-", "Manage Guild", "Dyno, Carl-bot, Sapphire", "Yes", "No",
     "Event groups: messages, members, roles, channels, voice, server, invites, moderation, automod. Each group can target its own channel."),
    ("?logging ignore", "Logging", "logignore", "Both",
     "Exclude channels, roles or members from logging.",
     "<add|remove|list> <channel|role|member>", "Manage Guild", "Carl-bot", "Yes", "No", "-"),
    ("?modlog", "Logging", "-", "Both",
     "Set the channel that receives numbered moderation cases.",
     "[channel]", "Manage Guild", "Dyno, Carl-bot", "Yes", "No", "-"),
    ("?snipe", "Logging", "-", "Both",
     "Show the most recently deleted message in a channel.",
     "[channel] [index]", "Manage Messages", "Carl-bot", "Yes", "No",
     "In-memory ring buffer, capped per channel; never persisted to disk."),
    ("?editsnipe", "Logging", "esnipe", "Both",
     "Show the before/after of the most recently edited message in a channel.",
     "[channel] [index]", "Manage Messages", "Carl-bot", "Yes", "No", "-"),
    ("?auditlog", "Logging", "audit", "Both",
     "Show recent Discord audit-log entries, filterable by action and actor.",
     "[action] [user] [limit=10]", "View Audit Log", "Dyno", "Yes", "No", "-"),
]

LEVELING: List[Row] = [
    ("?rank", "Leveling & Economy", "level, xp", "Both",
     "Show a member's level, XP, progress bar and server position as a rendered card.",
     "[user]", "None", "Carl-bot, Lawliet, Sapphire", "No", "No",
     "Card rendered locally with Pillow; falls back to a plain embed if image generation fails."),
    ("?leaderboard", "Leveling & Economy", "lb, top", "Both",
     "Paginated XP leaderboard for the server.",
     "[page]", "None", "Carl-bot, Lawliet, Sapphire", "Yes", "No", "-"),
    ("?levelconfig", "Leveling & Economy", "-", "Both",
     "Open the leveling panel: XP rate, cooldown, announcement channel and message, role stacking, blacklists.",
     "-", "Manage Guild", "Carl-bot, Sapphire", "Yes", "Yes", "-"),
    ("?levelrole", "Leveling & Economy", "-", "Both",
     "Manage roles awarded at a given level.",
     "<add|remove|list> [level] [role]", "Manage Guild", "Carl-bot, Sapphire", "Yes", "No",
     "Stack or replace behaviour is a toggle in ?levelconfig."),
    ("?xpblacklist", "Leveling & Economy", "-", "Both",
     "Exclude channels or roles from earning XP.",
     "<add|remove|list> <channel|role>", "Manage Guild", "Carl-bot", "Yes", "No", "-"),
    ("?setxp", "Leveling & Economy", "givexp", "Both",
     "Set, add or subtract a member's XP.",
     "<user> <set|add|remove> <amount>", "Manage Guild", "Carl-bot, Sapphire", "No", "No", "-"),
    ("?resetxp", "Leveling & Economy", "-", "Both",
     "Reset XP for one member or the whole server.",
     "<user|all>", "Administrator", "Carl-bot", "Yes", "No", "Confirmation buttons on 'all'."),
    ("?voicexp", "Leveling & Economy", "-", "Both",
     "Toggle XP earned for time spent in voice channels, and set the per-minute rate.",
     "<on|off|rate> [amount]", "Manage Guild", "Lawliet", "Yes", "Yes",
     "Ignores AFK channel and self-deafened members."),
    ("?balance", "Leveling & Economy", "bal, coins", "Both",
     "Show a member's currency balance.", "[user]", "None", "Lawliet", "No", "No", "-"),
    ("?daily", "Leveling & Economy", "-", "Both",
     "Claim a daily currency reward with a streak bonus.", "-", "None", "Lawliet", "No", "No",
     "24h cooldown enforced in the DB, not in memory, so it survives restarts."),
    ("?work", "Leveling & Economy", "-", "Both",
     "Earn currency on a cooldown from a randomised job table.", "-", "None", "Lawliet", "No", "No", "-"),
    ("?pay", "Leveling & Economy", "give", "Both",
     "Transfer currency to another member.", "<user> <amount>", "None", "Lawliet", "Yes", "No",
     "Confirmation buttons above a configurable amount."),
    ("?shop", "Leveling & Economy", "-", "Both",
     "Browse purchasable roles and items configured for the server.", "[page]", "None", "Lawliet", "Yes", "No", "-"),
    ("?buy", "Leveling & Economy", "-", "Both",
     "Purchase a shop item with currency.", "<item>", "None", "Lawliet", "Yes", "No", "-"),
    ("?inventory", "Leveling & Economy", "inv", "Both",
     "List items a member owns.", "[user]", "None", "Lawliet", "Yes", "No", "-"),
    ("?shopconfig", "Leveling & Economy", "-", "Both",
     "Add, edit, price or remove shop items and role rewards.",
     "<add|edit|remove|list>", "Manage Guild", "Lawliet", "Yes", "Yes", "-"),
    ("?gamble", "Leveling & Economy", "slots, bet", "Both",
     "Wager currency on a simple chance game.", "<amount>", "None", "Lawliet", "Yes", "No",
     "Payout table and max bet configurable; can be disabled per guild."),
    ("?ecoadmin", "Leveling & Economy", "-", "Both",
     "Add, remove or reset a member's currency, and configure earn rates.",
     "<add|remove|set|reset> <user> [amount]", "Administrator", "Lawliet", "Yes", "No", "-"),
    ("?importlevels", "Leveling & Economy", "mee6import", "Both",
     "Import an existing leveling table from another bot.",
     "<mee6|csv>", "Administrator", "Carl-bot", "Yes", "No",
     "EXTERNAL API: the Mee6 leaderboard endpoint is public but unofficial and often rate-limited or disabled. CSV import is the reliable path and is always available."),
]

UTILITY: List[Row] = [
    ("?serverinfo", "Utility", "guildinfo, si", "Both",
     "Show server creation date, owner, counts, boost tier, features and icon.",
     "-", "None", "Dyno, Carl-bot, Lawliet, Sapphire", "No", "No", "-"),
    ("?userinfo", "Utility", "whois, ui", "Both",
     "Show a member's join date, account age, roles, permissions and moderation summary.",
     "[user]", "None", "Dyno, Carl-bot, Lawliet, Sapphire", "Yes", "No",
     "Buttons jump to that member's warnings and mod-log history for staff."),
    ("?avatar", "Utility", "av, pfp", "Both",
     "Show a member's avatar, with server-specific avatar and format links.",
     "[user]", "None", "Dyno, Carl-bot, Lawliet", "Yes", "No", "-"),
    ("?banner", "Utility", "-", "Both",
     "Show a member's or the server's banner.", "[user]", "None", "Lawliet", "No", "No", "-"),
    ("?channelinfo", "Utility", "-", "Both",
     "Show a channel's ID, type, topic, slowmode, category and creation date.",
     "[channel]", "None", "Dyno", "No", "No", "-"),
    ("?membercount", "Utility", "mc", "Both",
     "Show member, human, bot and online counts.", "-", "None", "Dyno", "No", "No", "-"),
    ("?ping", "Utility", "-", "Both",
     "Show gateway and API round-trip latency, plus DB query time.", "-", "None",
     "Dyno, Carl-bot, Sapphire", "No", "No", "-"),
    ("?botinfo", "Utility", "about, stats, uptime", "Both",
     "Show bot uptime, version, discord.py version, memory use and command counts.",
     "-", "None", "Dyno, Carl-bot, Sapphire", "No", "No", "-"),
    ("?help", "Utility", "h, commands", "Both",
     "Interactive help: category select menu, per-command detail, search.",
     "[command]", "None", "Dyno, Carl-bot, Lawliet, Sapphire", "Yes", "Yes",
     "Modal used for the search box."),
    ("?emojis", "Utility", "emotes", "Both",
     "List the server's custom emoji with IDs.", "[search]", "None", "Dyno, Carl-bot", "Yes", "No", "Paginated."),
    ("?addemoji", "Utility", "steal, addemote", "Both",
     "Add an emoji from an image, a URL, or by copying one from another message.",
     "<name> [url|attachment|emoji]", "Manage Expressions", "Dyno, Carl-bot", "No", "No", "-"),
    ("?delemoji", "Utility", "-", "Both",
     "Delete a custom emoji.", "<emoji>", "Manage Expressions", "Dyno", "No", "No", "-"),
    ("?embed", "Utility", "-", "Both",
     "Build, post and later edit rich embeds through a guided builder, or from raw JSON.",
     "<create|edit|source|json> [args]", "Manage Messages", "Carl-bot, Sapphire", "Yes", "Yes",
     "Builder uses buttons for each field and modals for text input, with live preview before posting."),
    ("?say", "Utility", "echo", "Both",
     "Post a plain message as the bot in a chosen channel.",
     "[channel] <message>", "Manage Messages", "Carl-bot, Sapphire", "No", "No",
     "Mentions are sanitised unless the invoker has Mention Everyone."),
    ("?announce", "Utility", "-", "Both",
     "Post an announcement with an optional role ping, honouring a configured announcement role.",
     "<channel> <message> [--mention <role>]", "Manage Guild", "Dyno", "Yes", "Yes", "-"),
    ("?poll", "Utility", "-", "Both",
     "Create a button-based poll with up to 10 options and live vote counts.",
     "<question> <option1> <option2> ... [--duration <time>] [--multi]",
     "Manage Messages", "Dyno, Carl-bot, Sapphire", "Yes", "Yes",
     "Persistent view; results survive restarts and are stored per voter to prevent double-voting."),
    ("?remind", "Utility", "remindme, timer", "Both",
     "Set a personal or channel reminder.", "<duration> <text>", "None",
     "Dyno, Carl-bot", "Yes", "No", "Cancel button on the confirmation message."),
    ("?reminders", "Utility", "-", "Both",
     "List and cancel your pending reminders.", "-", "None", "Dyno, Carl-bot", "Yes", "No", "-"),
    ("?highlight", "Utility", "hl", "Both",
     "Manage personal keyword highlights that DM you when someone says them.",
     "<add|remove|list|block|unblock|clear> [word]", "None", "Dyno, Carl-bot", "Yes", "No",
     "Per-user, not per-guild. Blocked channels and members are respected."),
    ("?afk", "Utility", "-", "Both",
     "Set an AFK status shown when someone mentions you; clears on your next message.",
     "[message]", "None", "Dyno", "No", "No", "-"),
    ("?starboard", "Utility", "star", "Both",
     "Configure the starboard: channel, star emoji, threshold, self-star, NSFW and channel blacklists.",
     "<channel|emoji|threshold|selfstar|blacklist|top> [args]", "Manage Guild",
     "Carl-bot", "Yes", "Yes", "?starboard top shows the most-starred messages of all time."),
    ("?suggest", "Utility", "-", "Both",
     "Submit a suggestion to the configured suggestions channel with vote buttons.",
     "<text>", "None", "Carl-bot", "Yes", "Yes", "Modal used for the suggestion body on the slash version."),
    ("?suggestion", "Utility", "-", "Both",
     "Approve, deny, implement or comment on a suggestion.",
     "<approve|deny|implement|comment> <id> [reason]", "Manage Guild", "Carl-bot", "Yes", "Yes", "-"),
    ("?firstmessage", "Utility", "first", "Both",
     "Link the first message ever sent in a channel.", "[channel]", "None", "Carl-bot", "No", "No", "-"),
    ("?color", "Utility", "colour", "Both",
     "Preview a hex or named colour, or generate a random one.", "[hex|random]", "None", "Dyno", "No", "No", "-"),
    ("?inviteinfo", "Utility", "-", "Both",
     "Show which server an invite code points to, its uses and expiry.",
     "<code>", "None", "Dyno", "No", "No", "-"),
    ("?timestamp", "Utility", "time", "Both",
     "Convert a date/time into every Discord timestamp format.",
     "<when> [timezone]", "None", "Deezee original", "No", "No",
     "Not in the source bots; trivial to build and heavily used by staff writing announcements."),
]

FUN: List[Row] = [
    ("?8ball", "Fun/Social", "eightball", "Both",
     "Answer a yes/no question at random.", "<question>", "None",
     "Carl-bot, Lawliet, Sapphire", "No", "No", "-"),
    ("?coinflip", "Fun/Social", "flip", "Both",
     "Flip a coin.", "-", "None", "Dyno, Carl-bot", "No", "No", "-"),
    ("?roll", "Fun/Social", "dice", "Both",
     "Roll dice in NdN notation with modifiers.", "[NdN=1d6]", "None", "Lawliet, Sapphire", "No", "No", "-"),
    ("?choose", "Fun/Social", "pick", "Both",
     "Pick one option from a comma-separated list.", "<a, b, c...>", "None",
     "Carl-bot, Sapphire", "No", "No", "-"),
    ("?rps", "Fun/Social", "-", "Both",
     "Play rock-paper-scissors against the bot or challenge another member.",
     "[opponent]", "None", "Lawliet", "Yes", "No", "Buttons for the three choices; both players get their own locked view."),
    ("?fancy", "Fun/Social", "aesthetic, smallcaps, fraktur, owofy, clap, emojify, reverse", "Both",
     "Transform text: fullwidth, small caps, fraktur, owo, clap-separated, emoji letters or reversed.",
     "<style> <text>", "None", "Carl-bot", "No", "No",
     "One command with a style argument, replacing Carl-bot's eight separate commands."),
    ("?cat", "Fun/Social", "-", "Both",
     "Post a random cat picture.", "-", "None", "Dyno", "No", "No",
     "EXTERNAL API: TheCatAPI, free and keyless. Degrades to a bundled local image set if the request fails."),
    ("?dog", "Fun/Social", "pug", "Both",
     "Post a random dog picture.", "-", "None", "Dyno", "No", "No",
     "EXTERNAL API: dog.ceo, free and keyless."),
    ("?urban", "Fun/Social", "ud", "Both",
     "Look up an Urban Dictionary definition.", "<term>", "None", "Carl-bot", "Yes", "No",
     "EXTERNAL API: Urban Dictionary public API, free and keyless. NSFW-gated to age-restricted channels."),
    ("?profile", "Fun/Social", "-", "Both",
     "Show a member's social profile: bio, reputation, level and currency.",
     "[user]", "None", "Lawliet", "Yes", "Yes", "Edit button opens a modal for the bio."),
    ("?rep", "Fun/Social", "-", "Both",
     "Give another member a reputation point, once every 24 hours.",
     "<user> [reason]", "None", "Lawliet", "No", "No", "-"),
]

GIVEAWAYS: List[Row] = [
    ("?giveaway", "Giveaways & Events", "g, gw", "Both",
     "Root giveaway command: start, end, reroll, cancel, edit and list giveaways.",
     "<start|end|reroll|cancel|edit|list> [args]", "Manage Guild",
     "Dyno, Sapphire, Lawliet", "Yes", "Yes",
     "Entry is a persistent button, re-registered on ready so running giveaways survive restarts."),
    ("?giveaway start", "Giveaways & Events", "gstart", "Both",
     "Launch a giveaway with duration, winner count, prize and entry requirements.",
     "<duration> <winners> <prize> [--role <role>] [--level <n>] [--messages <n>]",
     "Manage Guild", "Dyno, Sapphire, Lawliet", "Yes", "Yes",
     "Slash version opens a setup modal. Requirements are checked at entry time and again at draw time."),
    ("?giveaway end", "Giveaways & Events", "gend", "Both",
     "End a running giveaway early and draw winners now.",
     "<message_id>", "Manage Guild", "Dyno, Sapphire", "Yes", "No", "-"),
    ("?giveaway reroll", "Giveaways & Events", "greroll", "Both",
     "Draw replacement winners for a finished giveaway.",
     "<message_id> [count=1]", "Manage Guild", "Dyno, Sapphire", "No", "No",
     "Excludes previous winners unless --allow-repeat is passed."),
    ("?giveaway cancel", "Giveaways & Events", "-", "Both",
     "Cancel a giveaway without drawing any winner.",
     "<message_id>", "Manage Guild", "Sapphire", "Yes", "No", "Confirmation buttons."),
    ("?giveaway list", "Giveaways & Events", "-", "Both",
     "List active giveaways with their end times and entry counts.",
     "-", "Manage Guild", "Sapphire", "Yes", "No", "-"),
    ("?schedule", "Giveaways & Events", "scheduled", "Both",
     "Schedule a message or announcement to post once at a future time, or on a repeating interval.",
     "<create|list|delete> [args]", "Manage Guild", "Carl-bot, Sapphire", "Yes", "Yes",
     "Covers Carl-bot's feeds/autofeeds and Sapphire's scheduled messages. Backed by the same scheduler as timed punishments."),
]

TAGS: List[Row] = [
    ("?tag", "Tags & Custom Commands", "t", "Both",
     "Show a stored tag, or manage tags with the subcommands below.",
     "<name> | <create|edit|delete|info|list|raw|search|alias|transfer> [args]",
     "None (Manage Messages to edit others')", "Dyno, Carl-bot", "Yes", "Yes",
     "Create/edit use a modal so multi-line tag bodies are easy to enter."),
    ("?tag create", "Tags & Custom Commands", "add", "Both",
     "Create a new tag owned by you.", "<name> <content>", "None", "Dyno, Carl-bot", "No", "Yes", "-"),
    ("?tag edit", "Tags & Custom Commands", "-", "Both",
     "Edit a tag you own.", "<name> <content>", "None (owner) / Manage Messages",
     "Dyno, Carl-bot", "No", "Yes", "-"),
    ("?tag delete", "Tags & Custom Commands", "remove", "Both",
     "Delete a tag you own.", "<name>", "None (owner) / Manage Messages",
     "Dyno, Carl-bot", "Yes", "No", "-"),
    ("?tag list", "Tags & Custom Commands", "-", "Both",
     "List tags, optionally filtered to one owner.", "[user] [page]", "None",
     "Dyno, Carl-bot", "Yes", "No", "Paginated."),
    ("?tag info", "Tags & Custom Commands", "-", "Both",
     "Show a tag's owner, use count and creation date.", "<name>", "None",
     "Dyno, Carl-bot", "No", "No", "-"),
    ("?tag raw", "Tags & Custom Commands", "-", "Both",
     "Show a tag's source with formatting escaped, for copying.",
     "<name>", "None", "Carl-bot", "No", "No", "-"),
    ("?tag search", "Tags & Custom Commands", "-", "Both",
     "Search tag names and content.", "<query>", "None", "Carl-bot", "Yes", "Yes", "-"),
    ("?tag alias", "Tags & Custom Commands", "-", "Both",
     "Point a second name at an existing tag.", "<alias> <target>", "None", "Carl-bot", "No", "No", "-"),
    ("?tag transfer", "Tags & Custom Commands", "claim", "Both",
     "Transfer tag ownership, or claim a tag whose owner left the server.",
     "<name> [user]", "None (owner) / Manage Messages", "Carl-bot", "Yes", "No", "-"),
    ("?customcommand", "Tags & Custom Commands", "cc", "Both",
     "Create prefix commands that reply with text or an embed and can add/remove roles.",
     "<add|edit|remove|list|show> [args]", "Manage Guild", "Dyno, Sapphire", "Yes", "Yes",
     "Distinct from tags: custom commands can carry role actions and permission restrictions; tags cannot."),
    ("?autoresponse", "Tags & Custom Commands", "ar, autoresponder", "Both",
     "Reply automatically to messages matching a trigger.",
     "<add|remove|list|ignore|unignore> [trigger] [response] [--mode contains|exact|startswith|endswith|regex]",
     "Manage Guild", "Carl-bot, Sapphire", "Yes", "Yes",
     "Regex mode is compiled with a timeout guard to prevent catastrophic backtracking."),
    ("?variables", "Tags & Custom Commands", "vars", "Both",
     "List every variable usable in tags, custom commands, autoresponses and welcome messages.",
     "-", "None", "Carl-bot", "Yes", "No",
     "Deezee uses a restricted variable engine ({user}, {server}, {channel}, {count}, {random:a|b}, {choose:...}) rather than a full scripting language, to avoid arbitrary execution."),
]

CONFIG: List[Row] = [
    ("?config", "Server Config", "settings, setup", "Both",
     "Master configuration panel. Every guild setting is reachable from here without a browser.",
     "-", "Manage Guild", "Deezee original (replaces all five dashboards)", "Yes", "Yes",
     "Paginated embed with one button per category, each opening a sub-panel of select menus, toggles and modals. This is the sole configuration surface: there is no web dashboard."),
    ("?prefix", "Server Config", "-", "Both",
     "Show or change the command prefix for this server.", "[new_prefix]", "Manage Guild",
     "Dyno, Carl-bot, Sapphire", "No", "No", "Mentioning the bot always works as a prefix, so a bad prefix can never lock you out."),
    ("?welcome", "Server Config", "greet, joinmessage", "Both",
     "Configure the join message: channel, text or embed, DM copy, image card, delete-after.",
     "<channel|message|embed|dm|test|off>", "Manage Guild",
     "Carl-bot, Sapphire, Dyno", "Yes", "Yes", "Test button posts a preview using the invoker as the fake joiner."),
    ("?goodbye", "Server Config", "farewell, leavemessage", "Both",
     "Configure the leave message: channel, text or embed, delete-after.",
     "<channel|message|embed|test|off>", "Manage Guild",
     "Carl-bot, Sapphire", "Yes", "Yes", "-"),
    ("?modroles", "Server Config", "-", "Both",
     "Define which roles count as moderators for permission checks.",
     "<add|remove|list> [role]", "Manage Guild", "Dyno", "Yes", "No",
     "Lets you grant moderation commands without granting Discord-level permissions."),
    ("?adminroles", "Server Config", "-", "Both",
     "Define which roles count as administrators for permission checks.",
     "<add|remove|list> [role]", "Administrator", "Dyno", "Yes", "No", "-"),
    ("?protectedroles", "Server Config", "immune", "Both",
     "Roles that can never be targeted by moderation or automod actions.",
     "<add|remove|list> [role]", "Manage Guild", "Dyno", "Yes", "No", "-"),
    ("?permissions", "Server Config", "perms", "Both",
     "Per-command overrides: allow or deny a command for a role, channel or member.",
     "<allow|deny|reset|show> <command> <role|channel|member>", "Administrator",
     "Sapphire, Dyno", "Yes", "No", "-"),
    ("?command", "Server Config", "-", "Both",
     "Enable or disable an individual command in this server.",
     "<enable|disable|list> <command>", "Manage Guild", "Dyno", "Yes", "No", "-"),
    ("?module", "Server Config", "modules, cog", "Both",
     "Enable or disable a whole feature module.",
     "<enable|disable|list> <module>", "Manage Guild", "Dyno", "Yes", "No", "-"),
    ("?ignore", "Server Config", "-", "Both",
     "Stop the bot responding to commands in a channel, from a role, or from a member.",
     "<add|remove|list> <channel|role|member>", "Manage Guild", "Dyno", "Yes", "No", "-"),
    ("?configexport", "Server Config", "backup", "Both",
     "Export this server's full configuration as a JSON file attachment.",
     "-", "Administrator", "Deezee original", "No", "No",
     "Replaces the dashboard's implicit backup. Secrets are never included."),
    ("?configimport", "Server Config", "restore", "Both",
     "Restore configuration from a previously exported JSON file.",
     "<attachment>", "Administrator", "Deezee original", "Yes", "No",
     "Confirmation buttons; shows a diff of what will change before applying."),
    ("?configreset", "Server Config", "-", "Both",
     "Reset one module's settings, or every setting, to defaults.",
     "<module|all>", "Administrator", "Deezee original", "Yes", "No",
     "Destructive: double confirmation, invoker-locked."),
    ("?import", "Server Config", "migrate", "Both",
     "Open the migration panel: pull data and settings off the bots Deezee is replacing, into a staging area that must be reviewed before anything goes live.",
     "-", "Administrator", "Deezee original", "Yes", "Yes",
     "Nothing is ever applied automatically and nothing is ever deleted. Every import writes to a draft table; ?import review shows a diff and only an explicit button press commits it. Deezee never touches the other bots, never removes them, and never edits their data - removing them stays a manual job for the server managers."),
    ("?import levels", "Server Config", "-", "Both",
     "Import a leveling leaderboard from another bot, mapping their XP curve onto Deezee's.",
     "<mee6|arcane|csv> [attachment]", "Administrator", "Deezee original", "Yes", "No",
     "mee6: public JSON endpoint, requires the server's leaderboard to be set public in their dashboard, 1000 members max. arcane: scrapes arcane.bot/lb/<guild_id>, Cloudflare-protected and capped at roughly 100 entries on their free tier. csv: universal fallback taking user_id,xp and always available. XP is converted through each bot's published level curve, not copied raw, so ranks land in the same place."),
    ("?import modlog", "Server Config", "-", "Both",
     "Rebuild moderation history by parsing an existing mod-log channel's embeds into Deezee cases.",
     "<channel> [limit=5000]", "Administrator", "Deezee original", "Yes", "No",
     "Works because the mod-log embeds already live in your server and any bot with Read Message History can read them. Recognises Dyno, Carl-bot and Sapphire case formats. Preserves original case numbers, action, target, moderator and reason; flags anything it cannot parse for manual review rather than guessing."),
    ("?import reactionroles", "Server Config", "-", "Both",
     "Adopt an existing reaction-role message: read its reactions, then map each emoji to a role.",
     "<message_link>", "Administrator", "Deezee original", "Yes", "No",
     "The emoji on the message are readable, but which role each grants is stored in the other bot's database and is not. A select menu per emoji is shown so the mapping is set once, by hand. Deezee can either take over the existing message or rebuild it as buttons."),
    ("?import capture", "Server Config", "-", "Both",
     "Listen in a channel while you run the other bots' own config commands, and parse whatever they print into a draft configuration.",
     "<channel> [duration=5m]", "Administrator", "Deezee original", "Yes", "No",
     "Dyno, Carl-bot and Sapphire have no public config API, but they will all print their current settings into chat on request. This command records what they post during the window and parses it. Understands welcome and goodbye messages, autoresponse lists, tag lists, automod word lists and logging channels. Everything captured lands in the draft, never live."),
    ("?import paste", "Server Config", "-", "Both",
     "Paste a setting's raw text straight into the draft when no automatic path exists.",
     "<setting>", "Administrator", "Deezee original", "Yes", "Yes",
     "Modal-based fallback for anything ?import capture cannot read. Covers welcome, goodbye, ban message, autoresponses, tags and banned-word lists. Variables from the source bot are translated to Deezee's equivalents where they map, and flagged where they do not."),
    ("?import review", "Server Config", "-", "Both",
     "Show a field-by-field diff of the staged draft against the live configuration, then apply or discard it.",
     "[section]", "Administrator", "Deezee original", "Yes", "No",
     "The only command that writes imported data into live config. Per-section apply, so you can accept the leveling import and reject the automod import. Every apply is reversible via ?import undo for 7 days."),
    ("?import status", "Server Config", "-", "Both",
     "Migration checklist: what has been imported, what is still pending, and what has no automatic path.",
     "-", "Administrator", "Deezee original", "Yes", "No",
     "Also lists which of the old bots are still present in the guild, as a reminder - it never removes them. Ends with an explicit 'safe to remove' verdict per bot once its data has been imported and applied."),
    ("?import undo", "Server Config", "-", "Both",
     "Roll back a previously applied import section to the configuration that preceded it.",
     "<section>", "Administrator", "Deezee original", "Yes", "No",
     "Snapshots are kept for 7 days after each apply."),
    ("?diagnose", "Server Config", "-", "Both",
     "Explain why a command or module is not working here: missing bot permissions, disabled module, ignored channel, role override.",
     "<command|module>", "Manage Guild", "Dyno", "No", "No",
     "Highest-value support command; keeps configuration debuggable without a dashboard."),
]

ALL_ROWS: List[Row] = (
    MODERATION + AUTOMOD + ANTIRAID + ROLES + LOGGING
    + LEVELING + UTILITY + FUN + GIVEAWAYS + TAGS + CONFIG
)

# Colour per category, used to tint the Category cell.
CATEGORY_FILLS = {
    "Moderation": "FFE2E2",
    "Automod": "FFEFD5",
    "Anti-raid & Verification": "FFF7C2",
    "Roles & Reaction Roles": "E2F0D9",
    "Logging": "DDEBF7",
    "Leveling & Economy": "E4DFEC",
    "Utility": "EDEDED",
    "Fun/Social": "FCE4EC",
    "Giveaways & Events": "E0F7F4",
    "Tags & Custom Commands": "FFF2CC",
    "Server Config": "D9E1F2",
}

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="2F3136")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMN_WIDTHS = [26, 24, 34, 18, 62, 46, 26, 30, 12, 12, 70]


def style_header(ws: Worksheet, headers: List[str]) -> None:
    """Write and style a header row, then freeze it."""
    ws.append(headers)
    for idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def build_commands_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Commands"
    style_header(ws, HEADERS)

    for row in ALL_ROWS:
        ws.append(list(row))

    for r in range(2, ws.max_row + 1):
        category = ws.cell(row=r, column=2).value
        fill_hex = CATEGORY_FILLS.get(str(category), "FFFFFF")
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if c == 2:
                cell.fill = PatternFill("solid", fgColor=fill_hex)
            if c == 1:
                cell.font = Font(name=FONT, size=10, bold=True)
        # Highlight rows that need an external service or are not being built.
        note = str(ws.cell(row=r, column=11).value or "")
        if note.startswith("NOT BUILT") or "EXTERNAL API REQUIRED" in note:
            ws.cell(row=r, column=11).font = Font(name=FONT, size=10, bold=True, color="C00000")

    for i, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"


def build_summary_sheet(wb: Workbook) -> None:
    """Category counts, computed with live formulas against the Commands sheet."""
    ws = wb.create_sheet("Summary")
    style_header(ws, ["Category", "Commands", "Uses buttons", "Uses modals", "Slash + prefix"])

    categories = list(CATEGORY_FILLS.keys())
    last = len(ALL_ROWS) + 1  # last data row on the Commands sheet

    for i, category in enumerate(categories, start=2):
        ws.cell(row=i, column=1, value=category)
        ws.cell(row=i, column=2, value=f'=COUNTIF(Commands!$B$2:$B${last},$A{i})')
        ws.cell(row=i, column=3,
                value=f'=COUNTIFS(Commands!$B$2:$B${last},$A{i},Commands!$I$2:$I${last},"Yes")')
        ws.cell(row=i, column=4,
                value=f'=COUNTIFS(Commands!$B$2:$B${last},$A{i},Commands!$J$2:$J${last},"Yes")')
        ws.cell(row=i, column=5,
                value=f'=COUNTIFS(Commands!$B$2:$B${last},$A{i},Commands!$D$2:$D${last},"Both")')

    total_row = len(categories) + 2
    ws.cell(row=total_row, column=1, value="TOTAL")
    for col in range(2, 6):
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col,
                value=f"=SUM({letter}2:{letter}{total_row - 1})")

    for r in range(2, total_row + 1):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10, bold=(r == total_row))
            cell.alignment = Alignment(vertical="center")
            cell.border = BORDER
            if c == 1 and r < total_row:
                cell.fill = PatternFill("solid", fgColor=CATEGORY_FILLS[str(cell.value)])

    for i, width in enumerate([28, 12, 14, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


NOTES_ROWS = [
    ("Prefix", "?", "Every command is registered as a prefix command. Mentioning the bot also works as a prefix, so a mistyped ?prefix can never lock the server out."),
    ("Slash coverage", "All commands",
     "Every command in this sheet is also a slash command. Commands whose arguments are long free text (tag bodies, embed content, ban lists) open a modal from the slash version instead of taking a long option string."),
    ("Deduplication", "5 bots -> 1",
     "Overlapping commands were collapsed. Example: Dyno ?ban, Carl-bot !ban, Lawliet L.ban and Sapphire s!ban all became a single ?ban; Carl-bot's eight separate text-transform commands became ?fancy <style>."),
    ("No dashboard", "Hard constraint",
     "There is no web UI of any kind. ?config is the only configuration surface, built from embeds, buttons, select menus and modals."),
    ("Persistent views", "bot.add_view()",
     "Reaction-role menus, giveaway entry buttons, verification buttons, poll buttons and suggestion vote buttons all use fixed custom_ids and are re-registered on ready, so they keep working after a restart."),
    ("Confirmation pattern", "Destructive actions",
     "ban, softban, massban, kick, purge >100, nuke, lockdown, massrole, clearwarns, clearnotes, resetxp all, configreset and configimport require button confirmation. The view is locked to the invoker and times out after 30 seconds."),
    ("EXTERNAL API - link scanning", "?scanlinks",
     "Google Safe Browsing v5 urls:search, free and non-commercial-only. Endpoint GET https://safebrowsing.googleapis.com/v5/urls:search?key=<KEY>&urls=<url>, repeat the urls parameter for up to 50 URLs, empty request body. IMPORTANT: v5 responds with application/x-protobuf only; alt=json is rejected with 400 Unsupported Output Format. The response message is tiny and fixed (threats[] of {url, threatTypes[]}, plus cacheDuration), so the bot ships a ~40-line hand-written varint decoder rather than taking a protobuf dependency. An empty threats list means clean. Verified working against this project's key on 2026-08-07. Only unknown links are sent: a bundled trusted-domain database plus a per-guild trusted list are checked first, and verdicts are cached, so the API sees a small fraction of posted links. Set GOOGLE_SAFE_BROWSING_KEY in .env. Without it, the trusted lists and the offline phishing blocklist still work. Do not build against v4: it is deprecated and shuts down 31 March 2027."),
    ("EXTERNAL API - fun content", "?cat ?dog ?urban",
     "TheCatAPI, dog.ceo and Urban Dictionary. All free and keyless; each has a graceful failure path."),
    ("EXTERNAL API - level import", "?importlevels mee6",
     "The Mee6 leaderboard endpoint is public but unofficial and frequently rate-limited. CSV import is the supported path."),
    ("NOT BUILT - IP / VPN / proxy detection", "Double Counter core feature",
     "Discord does not expose member IP addresses to bots. Double Counter obtains them by redirecting every joining member to its own hosted web page and fingerprinting the browser. Reproducing that means hosting a public website, which the no-dashboard constraint rules out. ?altcheck implements the strongest heuristics available without IP data."),
    ("NOT BUILT - music", "Sapphire beta feature",
     "Out of the stated scope (moderation, automod, roles, logging, leveling, utility). Requires Lavalink or a voice-audio stack and carries platform ToS risk. Say the word and it can be added as a separate cog."),
    ("Migration - never destructive", "?import",
     "Deezee never deletes, disables, edits or removes any other bot, and never touches their data. Every import is read-only at the source and writes to a staging table. Applying a draft requires an explicit button press in ?import review, is done per section, and is reversible for 7 days via ?import undo. Removing Dyno, Carl-bot, Lawliet, Sapphire and Double Counter from the server remains entirely a manual decision for the server managers."),
    ("Migration - what can be pulled automatically", "?import",
     "Mee6 XP via their public JSON endpoint; Arcane XP by scraping arcane.bot/lb/<guild_id>; any bot's XP via CSV; moderation history by parsing an existing mod-log channel's embeds; starboard seed data from an existing starboard channel; reaction-role messages, with the emoji-to-role mapping supplied by hand."),
    ("Migration - what cannot be pulled automatically", "?import capture / ?import paste",
     "Welcome and goodbye text, automod filters and banned-word lists, autoresponders, tags, custom commands, logging configuration, level roles, XP rates and all Double Counter settings live in the other bots' private databases behind authenticated dashboards, with no public API. Deezee will not log into those dashboards. Two supported paths instead: ?import capture parses the settings those bots print into chat when asked, and ?import paste takes the text by hand through a modal."),
    ("Hosting target", "Wispbyte",
     "Pterodactyl-style panel: no systemd, no Docker build step, no root. Dependencies install from requirements.txt at boot; SQLite lives on the persistent volume; the process must survive container restarts, which is why all schedulers rehydrate from the database on ready."),
    ("Scope", "Single server, per-guild schema",
     "Built for one guild, but every table is keyed by guild_id so a second server costs no migration."),
]


def build_notes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Notes & Flags")
    style_header(ws, ["Topic", "Applies to", "Detail"])

    for topic, applies, detail in NOTES_ROWS:
        ws.append([topic, applies, detail])

    for r in range(2, ws.max_row + 1):
        flagged = str(ws.cell(row=r, column=1).value).startswith(("NOT BUILT", "EXTERNAL API"))
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT, size=10, bold=(c == 1),
                             color="C00000" if flagged else "000000")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    for i, width in enumerate([34, 30, 110], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def main() -> None:
    wb = Workbook()
    build_commands_sheet(wb)
    build_summary_sheet(wb)
    build_notes_sheet(wb)

    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Deezee_Command_Sheet.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path} with {len(ALL_ROWS)} commands.")


if __name__ == "__main__":
    main()
